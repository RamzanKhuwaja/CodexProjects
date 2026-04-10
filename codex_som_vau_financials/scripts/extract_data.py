"""
extract_data.py — som_vau_financials
=====================================
Reads all 4 QuickBooks Excel source files once and writes
data/extracted/run_data.json as the single source of truth
for all 4 report scripts.

Run from project root:
    python scripts/extract_data.py
"""

import csv
import glob
import json
import os
import re
from datetime import datetime

import openpyxl
from project_context import (
    fiscal_year_bounds,
    fiscal_year_label,
    load_historical_context,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = os.path.join(BASE_DIR, "data", "current")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "extracted")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "run_data.json")

# Keywords for glob-based file detection (tolerates any filename prefix)
KW_PL_COMPARE = "Profit and Loss - Compare YTD"
KW_PL_3YR     = "Profit and Loss - Aug 2022"
KW_SH_CURRENT = "Shareholder advances - this fiscal year"
KW_SH_ALL     = "Shareholder Advances - all dates"

# ---------------------------------------------------------------------------
# Stable business rules
# ---------------------------------------------------------------------------
MARKETING_OBLIGATION_RATE = 0.03            # 3% of gross revenue

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_file(keyword, exts=(".xlsx", ".csv")):
    """Find a file in DATA_DIR whose name contains keyword (any prefix). Returns full path."""
    for ext in exts:
        matches = glob.glob(os.path.join(DATA_DIR, f"*{keyword}*{ext}"))
        if matches:
            return matches[0]
    raise FileNotFoundError(f"No file matching '*{keyword}*' found in {DATA_DIR}")


def safe_float(val):
    """Convert a cell value to float; handles comma and dollar-formatted CSV strings."""
    if val is None:
        return 0.0
    if isinstance(val, str):
        val = val.replace(",", "").replace("$", "").replace("\u2019", "").strip()
        if not val:
            return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def load_xlsx_rows(path, sheet_name=None):
    """Load workbook and return list of non-empty rows (values only)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(c is not None for c in row):
            rows.append(row)
    return rows


def load_csv_rows(path):
    """Load a CSV and return rows as tuples matching openpyxl format (None for empty cells)."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            converted = tuple(cell if cell.strip() else None for cell in row)
            if any(c is not None for c in converted):
                rows.append(converted)
    return rows


def load_file(path):
    """Dispatch to CSV or XLSX loader based on file extension."""
    if path.lower().endswith(".csv"):
        return load_csv_rows(path)
    return load_xlsx_rows(path)


def row_label(row):
    """Return the first non-None string in a row (the account label column).
    Strip both regular whitespace and non-breaking spaces (\xa0)."""
    for cell in row:
        if isinstance(cell, str):
            return cell.strip().rstrip('\xa0').strip()
    return ""


def change_pct(current, prior):
    """Percent change from prior to current; None if prior is 0."""
    if prior == 0.0:
        return None
    return round((current - prior) / abs(prior) * 100, 2)


# ---------------------------------------------------------------------------
# Step 1: Read the YTD Compare P&L
# ---------------------------------------------------------------------------

def read_pl_compare(path):
    print(f"Reading P&L Compare YTD...  ({os.path.basename(path)})")
    rows = load_file(path)

    # --- Header: extract cutoff date from rows 1-5 ---
    ytd_cutoff_date = "Unknown"
    for row in rows[:6]:
        label = row_label(row)
        m = re.search(r'(\w+ \d+,?\s*\d{4})-(\w+ \d+,?\s*\d{4})', label)
        if m:
            ytd_cutoff_date = m.group(2).strip()
            break
        for cell in row:
            if isinstance(cell, str) and " - " in cell and ("2026" in cell or "2025" in cell):
                parts = cell.split(" - ")
                if len(parts) == 2:
                    end = parts[1].replace("(PY YTD)", "").replace("(custom)", "").strip()
                    ytd_cutoff_date = end
                    break

    # --- Build a lookup: label → {cy, py} ---
    # Columns: 0=label, 1=CY YTD, 2=PY YTD, 3=PY-1 (custom)
    data = {}

    def store(label, row):
        cy = safe_float(row[1]) if len(row) > 1 else 0.0
        py = safe_float(row[2]) if len(row) > 2 else 0.0
        data[label] = {"cy": cy, "py": py}

    for row in rows:
        label = row_label(row)
        if label:
            store(label, row)

    def get(label):
        return data.get(label, {"cy": 0.0, "py": 0.0})

    ytd_tuition_current = get("4100 Tuition")["cy"]
    ytd_tuition_prior   = get("4100 Tuition")["py"]

    qb_profit           = get("Profit")["cy"]
    qb_profit_prior     = get("Profit")["py"]
    taxes_paid_8500     = get("6935 Corporate Tax Expense")["cy"]
    taxes_paid_8500_prior = get("6935 Corporate Tax Expense")["py"]
    canada_carbon_rebate = get("4110.1 Canada Carbon Rebate")["cy"]
    canada_carbon_rebate_prior = get("4110.1 Canada Carbon Rebate")["py"]

    # Marketing accounts (6200-series) — VAU uses 6207 Online Ads (not 6208)
    marketing_account_keys = [
        "6200 Marketing/Advertising/Promotion",
        "6201 Marketing Material",
        "6202 Space Rental",
        "6203 Mailing",
        "6204 Mobile Sign",
        "6205 Public Sign",
        "6206 Print Media",
        "6207 Online Ads",
        "6201.2 FTC",
        "6209 Event Sponsorship",
    ]
    marketing_accounts = {}
    for key in marketing_account_keys:
        d = get(key)
        if d["cy"] != 0.0 or d["py"] != 0.0:
            marketing_accounts[key] = {"current": d["cy"], "prior": d["py"]}

    total_marketing_cy = get("Total for 6200 Marketing/Advertising/Promotion")["cy"]
    total_marketing_py = get("Total for 6200 Marketing/Advertising/Promotion")["py"]

    # Key expense accounts
    expense_account_keys = [
        ("5780 Student Handouts", "5780"),
        ("6600 Insurance", "6600"),
        ("5710 Royalty fee", "5710"),
        ("5711 Service Fee", "5711"),
        ("Total for 6405 IT Expense", "6405_total"),
        ("6405 IT Expense", "6405"),
        ("6405.6 SoM IT Charges", "6405_6"),
        ("6406 Software Expense", "6406"),
        ("6405.1 Hardware", "6405_1"),
        ("Total for 5200 Payroll Expense", "5200_total"),
        ("5211 Wages", "5211"),
        ("5212 CPP Expense", "5212"),
        ("5213 EI Expense", "5213"),
        ("5214 EHT Expense", "5214"),
        ("5215 Employee Benefits", "5215"),
        ("5217 WSIB Expense", "5217"),
        ("5200 Payroll Expense", "5200"),
        ("5100 Materials and Supplies", "5100"),
        ("6427 FirstData - Merchant Services", "6427"),
        ("6715 Repairs and maintenance", "6715"),
        ("5300 Teacher and CAA training", "5300"),
        ("5600 Campus Rent", "5600"),
        ("5605 Keele Street Campus Rent", "5605"),
        ("Total for 5600 Campus Rent", "5600_total"),
        ("5785 Teaching Supplies", "5785"),
        ("6250 Amortization/Depreciation expense", "6250"),
        ("6255 Amortization of franchise fee", "6255"),
        ("6111 Ceridian", "6111"),
        ("6116 Accounting Fees", "6116"),
        ("Total for 6110 Professional Fees", "6110_total"),
        ("6230 Competition Expense", "6230"),
        ("6260 Donations", "6260"),
        ("Total for 6300 Automobile costs", "6300_total"),
        ("Total for 6401 Office/Campus Expenses", "6401_total"),
        ("6409 SMS Conference", "6409"),
        ("Total for 6420 Finance charges", "6420_total"),
        ("6421 Bank charges", "6421"),
        ("6423 Interest expense - bank", "6423"),
        ("6425 CC charges", "6425"),
        ("6602 Cleaning Service", "6602"),
        ("6710 Utilities", "6710"),
        ("4900 Interest Income", "4900"),
        ("4921 Tuition Refunds", "4921"),
        ("Total for Income", "income_total"),
        ("Gross Profit", "gross_profit"),
        ("Total for Expenses", "expenses_total"),
    ]

    expenses = {}
    for full_label, code in expense_account_keys:
        d = get(full_label)
        if d["cy"] != 0.0 or d["py"] != 0.0:
            expenses[code] = {
                "label": full_label,
                "current_ytd": d["cy"],
                "prior_ytd": d["py"],
                "change_pct": change_pct(d["cy"], d["py"]),
            }

    return {
        "ytd_cutoff_date": ytd_cutoff_date,
        "ytd_tuition_current": ytd_tuition_current,
        "ytd_tuition_prior": ytd_tuition_prior,
        "qb_profit": qb_profit,
        "qb_profit_prior": qb_profit_prior,
        "taxes_paid_8500": taxes_paid_8500,
        "taxes_paid_8500_prior": taxes_paid_8500_prior,
        "canada_carbon_rebate": canada_carbon_rebate,
        "canada_carbon_rebate_prior": canada_carbon_rebate_prior,
        "marketing_accounts": marketing_accounts,
        "total_marketing_cy": total_marketing_cy,
        "total_marketing_py": total_marketing_py,
        "expenses": expenses,
    }


# ---------------------------------------------------------------------------
# Step 2: Read the 3-year aggregate P&L
# ---------------------------------------------------------------------------

def read_pl_3yr(path):
    print(f"Reading 3-year aggregate P&L...  ({os.path.basename(path)})")
    rows = load_file(path)

    data = {}
    for row in rows:
        label = row_label(row)
        if label:
            val = safe_float(row[1]) if len(row) > 1 else 0.0
            data[label] = val

    def get(label):
        return data.get(label, 0.0)

    tuition_total      = get("4100 Tuition")
    marketing_total    = get("Total for 6200 Marketing/Advertising/Promotion")
    ftc_total          = get("6201.2 FTC")
    it_total           = get("Total for 6405 IT Expense")
    payroll_total      = get("Total for 5200 Payroll Expense")
    handouts_total     = get("5780 Student Handouts")
    amortization_total = get("6250 Amortization/Depreciation expense") + get("6255 Amortization of franchise fee")

    return {
        "tuition_total":           tuition_total,
        "tuition_annual_avg":      round(tuition_total / 3, 2),
        "marketing_total":         marketing_total,
        "marketing_annual_avg":    round(marketing_total / 3, 2),
        "ftc_total":               ftc_total,
        "ftc_annual_avg":          round(ftc_total / 3, 2),
        "it_total":                it_total,
        "it_annual_avg":           round(it_total / 3, 2),
        "payroll_total":           payroll_total,
        "payroll_annual_avg":      round(payroll_total / 3, 2),
        "handouts_total":          handouts_total,
        "handouts_annual_avg":     round(handouts_total / 3, 2),
        "amortization_total":      amortization_total,
        "amortization_annual_avg": round(amortization_total / 3, 2),
    }


# ---------------------------------------------------------------------------
# Step 3: Read the shareholder files
# ---------------------------------------------------------------------------

def parse_shareholder_sheet(rows, person_label, acct_key):
    """
    Extract opening balance, closing balance, and all transactions
    for a given sub-account (e.g. '2901 Ramzan Khuwaja').

    Row structure (QuickBooks export):
        col 0: None or sub-account label or section label
        col 1: account name or 'Beginning Balance' or 'TOTAL'
        col 2: transaction date
        col 3: transaction type
        col 4: #
        col 5: Name
        col 6: Memo/Description
        col 7: Account full name
        col 8: Cleared
        col 9: Amount
        col 10: Balance
    """
    in_section = False
    opening_balance = 0.0
    closing_balance = 0.0
    transactions = []

    for row in rows:
        if row[0] == person_label:
            in_section = True
            continue

        if not in_section:
            continue

        label0 = str(row[0]) if row[0] is not None else ""
        label1 = str(row[1]) if row[1] is not None else ""

        if row[0] is not None and row[0] != person_label and label0.startswith("29"):
            break
        if label0.startswith("Total for 2900") or label0 == "TOTAL" or label1 == "TOTAL":
            break

        if label1 == "Beginning Balance":
            opening_balance = safe_float(row[10])
            continue

        if label0.startswith(f"Total for {person_label}"):
            break

        if label1 == acct_key and row[2] is not None:
            date_str  = str(row[2]) if row[2] is not None else ""
            memo      = str(row[6]) if row[6] is not None else ""
            name      = str(row[5]) if row[5] is not None else ""
            tx_type   = str(row[3]) if row[3] is not None else ""
            tx_num    = str(row[4]) if row[4] is not None else ""
            amount    = safe_float(row[9])
            balance   = safe_float(row[10])
            closing_balance = balance

            transactions.append({
                "date":    date_str,
                "type":    tx_type,
                "num":     tx_num,
                "name":    name,
                "memo":    memo,
                "amount":  round(amount, 2),
                "balance": round(balance, 2),
            })

    return {
        "opening_balance": round(opening_balance, 2),
        "closing_balance": round(closing_balance, 2),
        "transactions":    transactions,
    }


def parse_parent_shareholder_account(rows):
    """
    Extract opening and closing balances for the parent 2900 account.
    This account holds the offset that makes the shareholder net tie to the
    reviewed financial statements.
    """
    in_section = False
    opening_balance = 0.0
    closing_balance = 0.0
    for row in rows:
        label0 = str(row[0]) if row[0] is not None else ""
        label1 = str(row[1]) if row[1] is not None else ""

        if row[0] == "2900 Shareholder's Advance":
            in_section = True
            continue

        if not in_section:
            continue

        if label0.startswith("2901 ") or label0.startswith("2902 "):
            break

        if label1 == "Beginning Balance":
            opening_balance = safe_float(row[10])
            closing_balance = opening_balance
            continue

        if label0.startswith("Total for 2900 Shareholder's Advance"):
            break

        if label1 == "2900 Shareholder's Advance" and row[2] is not None:
            closing_balance = safe_float(row[10])

    return {
        "opening_balance": round(opening_balance, 2),
        "closing_balance": round(closing_balance, 2),
    }


def read_shareholder_file(path, label):
    print(f"Reading {label}...  ({os.path.basename(path)})")
    rows = load_file(path)

    parent = parse_parent_shareholder_account(rows)

    # VAU shareholders: Ramzan (2901) and Farah (2902)
    ramzan = parse_shareholder_sheet(rows, "2901 Ramzan Khuwaja", "2901 Ramzan Khuwaja")
    farah  = parse_shareholder_sheet(rows, "2902 Farah Khuwaja",   "2902 Farah Khuwaja")

    return parent, ramzan, farah


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("som_vau_financials — extract_data.py")
    print(f"Run at: {datetime.now().isoformat()}")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    pl_compare_path = find_file(KW_PL_COMPARE)
    pl_3yr_path     = find_file(KW_PL_3YR)
    sh_current_path = find_file(KW_SH_CURRENT)
    sh_all_path     = find_file(KW_SH_ALL)

    print(f"  P&L Compare   : {os.path.basename(pl_compare_path)}")
    print(f"  P&L 3-year    : {os.path.basename(pl_3yr_path)}")
    print(f"  SH Current    : {os.path.basename(sh_current_path)}")
    print(f"  SH All dates  : {os.path.basename(sh_all_path)}")

    # -----------------------------------------------------------------------
    # 1. YTD Compare P&L
    # -----------------------------------------------------------------------
    pl = read_pl_compare(pl_compare_path)
    historical = load_historical_context(BASE_DIR)
    prior_full_year_tuition = historical["prior_year"]["review_fs"]["full_year_tuition"]

    ytd_cutoff_date      = pl["ytd_cutoff_date"]
    ytd_tuition_current  = pl["ytd_tuition_current"]
    ytd_tuition_prior    = pl["ytd_tuition_prior"]
    qb_profit            = pl["qb_profit"]
    qb_profit_prior      = pl["qb_profit_prior"]
    taxes_paid_8500      = pl["taxes_paid_8500"]
    taxes_paid_8500_prior = pl["taxes_paid_8500_prior"]
    canada_carbon_rebate = pl["canada_carbon_rebate"]
    canada_carbon_rebate_prior = pl["canada_carbon_rebate_prior"]

    # -----------------------------------------------------------------------
    # 2. 3-year benchmarks
    # -----------------------------------------------------------------------
    benchmarks = read_pl_3yr(pl_3yr_path)

    # -----------------------------------------------------------------------
    # 3. Shareholder files
    # -----------------------------------------------------------------------
    sh_cy_parent, sh_cy_ramzan, sh_cy_farah = read_shareholder_file(sh_current_path, "Current-Year Shareholder")
    sh_all_parent, sh_all_ramzan, sh_all_farah = read_shareholder_file(sh_all_path, "All-Dates Shareholder")

    # -----------------------------------------------------------------------
    # 4. Derived / computed fields
    # -----------------------------------------------------------------------

    # Pre-tax income proxy: QB profit + taxes booked - non-taxable carbon rebate
    h1_pretax_proxy = qb_profit + taxes_paid_8500 - canada_carbon_rebate

    # YTD-to-annual ratio: use prior year YTD vs prior full year from reviewed statements
    ytd_to_annual_ratio = (ytd_tuition_prior / prior_full_year_tuition) if prior_full_year_tuition else None

    # Projected full year revenue (current year)
    projected_full_year = (ytd_tuition_current / ytd_to_annual_ratio) if ytd_to_annual_ratio else None

    # YoY growth %
    yoy_growth_pct = change_pct(ytd_tuition_current, ytd_tuition_prior)

    # Marketing obligations (3%)
    obligation_ytd       = round(ytd_tuition_current * MARKETING_OBLIGATION_RATE, 2)
    obligation_projected = round(projected_full_year * MARKETING_OBLIGATION_RATE, 2) if projected_full_year else None
    gap_ytd              = round(obligation_ytd - pl["total_marketing_cy"], 2)
    gap_projected        = round(obligation_projected - pl["total_marketing_cy"], 2) if obligation_projected else None

    # Shareholder balances (VAU sign convention: negative = shareholder OWES corp)
    ramzan_close = sh_cy_ramzan["closing_balance"]
    farah_cy_close = sh_cy_farah["closing_balance"]
    if abs(farah_cy_close) < 0.01:
        farah_cy_close = 0.00
    farah_all_close = sh_all_farah["closing_balance"]
    if abs(farah_all_close) < 0.01:
        farah_all_close = 0.00

    combined_closing = round(ramzan_close + farah_cy_close, 2)
    net_opening_balance = round(
        sh_cy_parent["opening_balance"] + sh_cy_ramzan["opening_balance"] + sh_cy_farah["opening_balance"],
        2,
    )
    net_current_balance = round(
        sh_all_parent["closing_balance"] + ramzan_close + farah_cy_close,
        2,
    )

    cutoff_dt = datetime.strptime(ytd_cutoff_date, "%B %d, %Y")
    fy_start, fy_end = fiscal_year_bounds(cutoff_dt.date())

    # -----------------------------------------------------------------------
    # 5. Build JSON
    # -----------------------------------------------------------------------
    run_data = {
        "meta": {
            "extracted_at":       datetime.now().isoformat(),
            "ytd_cutoff_date":    ytd_cutoff_date,
            "fiscal_year_start":  fy_start.isoformat(),
            "fiscal_year_end":    fy_end.isoformat(),
            "fiscal_year_label":  fiscal_year_label(fy_start, fy_end),
            "fy_label":           f"FY{fy_start.year}-{str(fy_end.year)[-2:]}",
        },
        "revenue": {
            "ytd_tuition_current":    round(ytd_tuition_current, 2),
            "ytd_tuition_prior_year": round(ytd_tuition_prior, 2),
            "full_year_prior":        round(prior_full_year_tuition, 2),
            "ytd_to_annual_ratio":    round(ytd_to_annual_ratio, 6) if ytd_to_annual_ratio else None,
            "projected_full_year":    round(projected_full_year, 2) if projected_full_year else None,
            "yoy_growth_pct":         yoy_growth_pct,
        },
        "income": {
            "qb_profit":             round(qb_profit, 2),
            "qb_profit_prior":       round(qb_profit_prior, 2),
            "taxes_paid_8500":       round(taxes_paid_8500, 2),
            "taxes_paid_8500_prior": round(taxes_paid_8500_prior, 2),
            "canada_carbon_rebate":  round(canada_carbon_rebate, 2),
            "canada_carbon_rebate_prior": round(canada_carbon_rebate_prior, 2),
            "h1_pretax_proxy":       round(h1_pretax_proxy, 2),
            "prior_ytd_pretax_proxy": round(qb_profit_prior + taxes_paid_8500_prior - canada_carbon_rebate_prior, 2),
        },
        "marketing": {
            "total_ytd_current":     round(pl["total_marketing_cy"], 2),
            "total_ytd_prior":       round(pl["total_marketing_py"], 2),
            "accounts":              {k: {"current": round(v["current"], 2), "prior": round(v["prior"], 2)}
                                      for k, v in pl["marketing_accounts"].items()},
            "obligation_ytd":        obligation_ytd,
            "obligation_projected":  obligation_projected,
            "gap_ytd":               gap_ytd,
            "gap_projected":         gap_projected,
        },
        "expenses": pl["expenses"],
        "benchmarks_3yr": {
            "tuition_total":           round(benchmarks["tuition_total"], 2),
            "tuition_annual_avg":      benchmarks["tuition_annual_avg"],
            "marketing_total":         round(benchmarks["marketing_total"], 2),
            "marketing_annual_avg":    benchmarks["marketing_annual_avg"],
            "ftc_total":               round(benchmarks["ftc_total"], 2),
            "ftc_annual_avg":          benchmarks["ftc_annual_avg"],
            "it_total":                round(benchmarks["it_total"], 2),
            "it_annual_avg":           benchmarks["it_annual_avg"],
            "payroll_total":           round(benchmarks["payroll_total"], 2),
            "payroll_annual_avg":      benchmarks["payroll_annual_avg"],
            "handouts_total":          round(benchmarks["handouts_total"], 2),
            "handouts_annual_avg":     benchmarks["handouts_annual_avg"],
            "amortization_total":      round(benchmarks["amortization_total"], 2),
            "amortization_annual_avg": benchmarks["amortization_annual_avg"],
        },
        "tax": {
            "historical_reference": {
                "prior_full_year_tuition": round(prior_full_year_tuition, 2),
                "prior_accounting_income_before_tax": round(
                    historical["prior_year"]["review_fs"]["net_income_before_tax"], 2
                ),
                "prior_current_income_taxes": round(
                    historical["prior_year"]["review_fs"]["current_income_taxes"] or 0.0, 2
                ),
                "prior_taxable_income": round(historical["prior_year"]["t2"]["taxable_income"], 2),
                "prior_net_income_for_tax": round(
                    historical["prior_year"]["t2"]["net_income_tax_purposes"] or 0.0, 2
                ),
                "prior_part_i_tax": round(historical["prior_year"]["t2"]["part_i_tax"] or 0.0, 2),
                "prior_total_tax": round(historical["prior_year"]["t2"]["total_tax_payable"], 2),
                "prior_balance_owing": round(historical["prior_year"]["t2"]["balance_owing"] or 0.0, 2),
                "sbd_limit": round(historical["sbd_limit"], 2),
            },
            "installments": [],
            "installment_status_source": "No installment payment status is derived from the provided QuickBooks files.",
        },
        "shareholder": {
            "parent_2900": {
                "opening_balance": sh_cy_parent["opening_balance"],
                "closing_balance": round(sh_all_parent["closing_balance"], 2),
                "current_year": {
                    "opening_balance": sh_cy_parent["opening_balance"],
                    "closing_balance": round(sh_cy_parent["closing_balance"], 2),
                },
                "all_dates": {
                    "opening_balance": sh_all_parent["opening_balance"],
                    "closing_balance": round(sh_all_parent["closing_balance"], 2),
                },
            },
            "ramzan": {
                "opening_balance": sh_cy_ramzan["opening_balance"],
                "closing_balance": round(ramzan_close, 2),
                "transactions":    sh_cy_ramzan["transactions"],
                "all_dates": {
                    "opening_balance": sh_all_ramzan["opening_balance"],
                    "closing_balance": round(sh_all_ramzan["closing_balance"], 2),
                    "transactions":    sh_all_ramzan["transactions"],
                },
            },
            "farah": {
                "opening_balance": sh_cy_farah["opening_balance"],
                "closing_balance": farah_cy_close,
                "transactions":    sh_cy_farah["transactions"],
                "all_dates": {
                    "opening_balance": sh_all_farah["opening_balance"],
                    "closing_balance": farah_all_close,
                    "transactions":    sh_all_farah["transactions"],
                },
            },
            "combined_closing": combined_closing,
            "net_opening_balance": net_opening_balance,
            "net_current_balance": net_current_balance,
        },
    }

    # -----------------------------------------------------------------------
    # 6. Write JSON
    # -----------------------------------------------------------------------
    print(f"\nWriting run_data.json...  ({OUTPUT_FILE})")
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(run_data, f, indent=2, ensure_ascii=False)
    print("Done.")

    # -----------------------------------------------------------------------
    # 7. Verification output
    # -----------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("VERIFICATION — key values")
    print("=" * 60)
    print(f"  YTD cutoff date          : {ytd_cutoff_date}")
    print(f"  YTD tuition (CY)         : ${ytd_tuition_current:,.2f}")
    print(f"  QB Profit                : ${qb_profit:,.2f}")
    print(f"  Marketing total          : ${pl['total_marketing_cy']:,.2f}")
    print(f"  Parent 2900 closing      : ${sh_all_parent['closing_balance']:,.2f}")
    print(f"  Ramzan closing balance   : ${ramzan_close:,.2f}")
    print(f"  Farah closing balance    : ${farah_cy_close:,.2f}")
    print(f"  Subaccount combined      : ${combined_closing:,.2f}")
    print(f"  Real net shareholder bal : ${net_current_balance:,.2f}")
    print(f"  Taxes paid (8500)        : ${taxes_paid_8500:,.2f}")
    print(f"  H1 pre-tax proxy         : ${h1_pretax_proxy:,.2f}")
    print(f"  Prior YTD pre-tax proxy  : ${qb_profit_prior + taxes_paid_8500_prior - canada_carbon_rebate_prior:,.2f}")
    if ytd_to_annual_ratio:
        print(f"  YTD/Annual ratio         : {ytd_to_annual_ratio:.4%}")
    if projected_full_year:
        print(f"  Projected full year rev  : ${projected_full_year:,.2f}")
    print(f"  Marketing obligation YTD : ${obligation_ytd:,.2f}")
    if obligation_projected:
        print(f"  Marketing obligation proj: ${obligation_projected:,.2f}")
    print(f"  Marketing gap (YTD)      : ${gap_ytd:,.2f}")
    if gap_projected:
        print(f"  Marketing gap (proj)     : ${gap_projected:,.2f}")
    print("=" * 60)
    print()


if __name__ == "__main__":
    main()
