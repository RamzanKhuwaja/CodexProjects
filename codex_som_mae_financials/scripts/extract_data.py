"""
extract_data.py — som_mae_financials
====================================
Reads all 4 QuickBooks Excel source files once and writes
data/extracted/run_data.json as the single source of truth
for all 4 report scripts.

Run from project root:
    python scripts/extract_data.py
"""

import csv
import glob
import json
import os
import re
from datetime import datetime

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = os.path.join(BASE_DIR, "data", "current")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "extracted")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "run_data.json")


def find_file(keyword, exts=(".xlsx", ".csv")):
    """Return the first file in DATA_DIR whose name contains `keyword`
    and has one of the given extensions. Tolerates any filename prefix.
    Raises FileNotFoundError if nothing matches."""
    for ext in exts:
        matches = glob.glob(os.path.join(DATA_DIR, f"*{keyword}*{ext}"))
        if matches:
            return matches[0]
    raise FileNotFoundError(
        f"No file matching '*{keyword}*' ({', '.join(exts)}) in {DATA_DIR}"
    )

# ---------------------------------------------------------------------------
# Known constants (from audited financials and T2 — do not recompute from QB)
# ---------------------------------------------------------------------------
PY_FULL_YEAR_TUITION = 3_020_723.00   # FY2024-25 audited
ROYALTY_RATE         = 0.12
MARKETING_OBLIGATION_RATE = 0.03

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_float(val):
    """Convert a cell value to float; treat None or empty as 0.0.
    Handles comma-formatted strings like '3,176,492.61' from CSV exports."""
    if val is None:
        return 0.0
    if isinstance(val, str):
        val = val.replace(',', '').replace('$', '').strip()
        if not val:
            return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def load_xlsx_rows(path, sheet_name="Sheet1"):
    """Load an XLSX workbook and return non-empty rows as tuples (values only)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(c is not None for c in row):
            rows.append(row)
    return rows


def load_csv_rows(path):
    """Load a CSV file and return non-empty rows as tuples.
    Empty cells become None to match openpyxl behaviour."""
    rows = []
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            converted = tuple(cell if cell.strip() else None for cell in row)
            if any(c is not None for c in converted):
                rows.append(converted)
    return rows


def load_file(path, sheet_name="Sheet1"):
    """Load either an XLSX or CSV file and return rows as tuples."""
    if path.lower().endswith('.csv'):
        return load_csv_rows(path)
    return load_xlsx_rows(path, sheet_name)


def row_label(row):
    """Return the first non-None string in a row (the account label column).
    Strip both regular whitespace and non-breaking spaces (\xa0)."""
    for cell in row:
        if isinstance(cell, str):
            return cell.strip().rstrip('\xa0').strip()
    return ""


def change_pct(current, prior):
    """Percent change from prior to current; None if prior is 0."""
    if prior == 0.0:
        return None
    return round((current - prior) / abs(prior) * 100, 2)


# ---------------------------------------------------------------------------
# Step 1: Read the YTD Compare P&L
# ---------------------------------------------------------------------------

def read_pl_compare(path):
    print(f"Reading P&L Compare YTD...  ({os.path.basename(path)})")
    rows = load_file(path)

    # --- Header: extract cutoff date from rows 1-5 ---
    # Row 3 (index 2) looks like: ('August 1, 2025-March 11, 2026', None, None, None)
    ytd_cutoff_date = "Unknown"
    for row in rows[:6]:
        label = row_label(row)
        # Match "August 1, 2025-March 11, 2026" style
        m = re.search(r'(\w+ \d+,?\s*\d{4})-(\w+ \d+,?\s*\d{4})', label)
        if m:
            ytd_cutoff_date = m.group(2).strip()
            break
        # Match the column header "Aug. 1 2025 - Mar. 11 2026" style
        for cell in row:
            if isinstance(cell, str) and " - " in cell and ("2026" in cell or "2025" in cell):
                # Try to get the end date
                parts = cell.split(" - ")
                if len(parts) == 2:
                    # Take the end portion and clean it
                    end = parts[1].replace("(PY YTD)", "").replace("(custom)", "").strip()
                    ytd_cutoff_date = end
                    break

    # --- Build a lookup: label → (col_index, row) for the data rows ---
    # Columns: 0=label, 1=CY YTD, 2=PY YTD, 3=PY-1 (custom)
    # We extract col 1 (current year) and col 2 (prior year)

    data = {}   # label → {"cy": float, "py": float}

    def store(label, row):
        cy = safe_float(row[1]) if len(row) > 1 else 0.0
        py = safe_float(row[2]) if len(row) > 2 else 0.0
        data[label] = {"cy": cy, "py": py}

    for row in rows:
        label = row_label(row)
        if label:
            store(label, row)

    # --- Extract specific accounts ---

    def get(label):
        return data.get(label, {"cy": 0.0, "py": 0.0})

    ytd_tuition_current = get("4100 Tuition")["cy"]
    ytd_tuition_prior   = get("4100 Tuition")["py"]

    qb_profit           = get("Profit")["cy"]
    taxes_paid_8500     = get("8500 Taxes Paid")["cy"]
    canada_carbon_rebate = get("Canada Carbon Rebate (Non-Taxable)")["cy"]

    # Marketing accounts (6200-series)
    marketing_accounts = {}
    marketing_account_keys = [
        "6200 Marketing/Advertising/Promotion",
        "62010 Marketing Wages",
        "6201 Marketing Material",
        "6202 Space Rental",
        "6203 Mailing",
        "6204 Mobile Sign",
        "6205 Public Sign",
        "6206 Print Media",
        "6208 Online",
        "6201.1 FTC",
        "6209 Event Sponsorship",
    ]
    for key in marketing_account_keys:
        d = get(key)
        if d["cy"] != 0.0 or d["py"] != 0.0:
            marketing_accounts[key] = {"current": d["cy"], "prior": d["py"]}

    # Total marketing = "Total for 6200 Marketing/Advertising/Promotion"
    total_marketing_cy = get("Total for 6200 Marketing/Advertising/Promotion")["cy"]
    total_marketing_py = get("Total for 6200 Marketing/Advertising/Promotion")["py"]

    # Key expense accounts
    expense_account_keys = [
        ("5780 Student Handouts", "5780"),
        ("6600 Insurance", "6600"),
        ("5710 Royalty fee", "5710"),
        ("Total for 6405 IT Expense", "6405_total"),
        ("6405 IT Expense", "6405"),
        ("6405.6 SoM IT Charges", "6405_6"),
        ("6406 Software Expense", "6406"),
        ("6405.1 Hardware", "6405_1"),
        ("Total for 5200 Payroll Expanse", "5200_total"),
        ("5211 Wages", "5211"),
        ("5212 CPP Expense", "5212"),
        ("5213 EI Expense", "5213"),
        ("5214 EHT Expense", "5214"),
        ("5215 Employee Benefits", "5215"),
        ("5217 WSIB Expense", "5217"),
        ("5219 FED.TAX", "5219"),
        ("5200 Payroll Expanse", "5200"),
        ("5100 Materials and Supplies", "5100"),
        ("6427 FirstData - Merchant Services", "6427"),
        ("6715 Repairs and maintenance", "6715"),
        ("5300 Teacher and CAA training", "5300"),
        ("5600 Campus Rent", "5600"),
        ("5606 #5 - 9293 Markham Road", "5606"),
        ("Total for 5600 Campus Rent", "5600_total"),
        ("5785 Teaching Supplies", "5785"),
        ("5786 Non-SoM Contests", "5786"),
        ("6250 Amortization/Depreciation expense", "6250"),
        ("6255 Amortization of franchise fee", "6255"),
        ("6111 Ceridian", "6111"),
        ("6112 Security Services", "6112"),
        ("6116 Accounting Fees", "6116"),
        ("Total for 6110 Professional Fees", "6110_total"),
        ("6230 Competition Expanse", "6230"),
        ("6260 Donations", "6260"),
        ("Total for 6300 Automobile costs", "6300_total"),
        ("Total for 6401 Office/Campus Expenses", "6401_total"),
        ("6409 SMS Conference", "6409"),
        ("Total for 6420 Finance charges", "6420_total"),
        ("6421 Bank charges", "6421"),
        ("6423 Interest expense - bank", "6423"),
        ("6425 CC charges", "6425"),
        ("6602 Cleaning Service", "6602"),
        ("6603 Floor Mats Service", "6603"),
        ("6615 Events Expense", "6615"),
        ("6630 Office Photocopying Expense", "6630"),
        ("6710 Utilities", "6710"),
        ("4900 Interest Income", "4900"),
        ("4921 Tuition Refunds", "4921"),
        ("Total for Income", "income_total"),
        ("Gross Profit", "gross_profit"),
        ("Total for Expenses", "expenses_total"),
    ]

    expenses = {}
    for full_label, code in expense_account_keys:
        d = get(full_label)
        if d["cy"] != 0.0 or d["py"] != 0.0:
            expenses[code] = {
                "label": full_label,
                "current_ytd": d["cy"],
                "prior_ytd": d["py"],
                "change_pct": change_pct(d["cy"], d["py"]),
            }

    return {
        "ytd_cutoff_date": ytd_cutoff_date,
        "ytd_tuition_current": ytd_tuition_current,
        "ytd_tuition_prior": ytd_tuition_prior,
        "qb_profit": qb_profit,
        "taxes_paid_8500": taxes_paid_8500,
        "canada_carbon_rebate": canada_carbon_rebate,
        "marketing_accounts": marketing_accounts,
        "total_marketing_cy": total_marketing_cy,
        "total_marketing_py": total_marketing_py,
        "expenses": expenses,
    }


# ---------------------------------------------------------------------------
# Step 2: Read the 3-year aggregate P&L
# ---------------------------------------------------------------------------

def read_pl_3yr(path):
    print(f"Reading 3-year aggregate P&L...  ({os.path.basename(path)})")
    rows = load_file(path)

    data = {}
    for row in rows:
        label = row_label(row)
        if label:
            val = safe_float(row[1]) if len(row) > 1 else 0.0
            data[label] = val

    def get(label):
        return data.get(label, 0.0)

    tuition_total      = get("4100 Tuition")
    marketing_total    = get("Total for 6200 Marketing/Advertising/Promotion")
    ftc_total          = get("6201.1 FTC")
    it_total           = get("Total for 6405 IT Expense")
    payroll_total      = get("Total for 5200 Payroll Expanse")
    amortization_total = get("6250 Amortization/Depreciation expense") + get("6255 Amortization of franchise fee")

    return {
        "tuition_total":        tuition_total,
        "tuition_annual_avg":   round(tuition_total / 3, 2),
        "marketing_total":      marketing_total,
        "marketing_annual_avg": round(marketing_total / 3, 2),
        "ftc_total":            ftc_total,
        "ftc_annual_avg":       round(ftc_total / 3, 2),
        "it_total":             it_total,
        "it_annual_avg":        round(it_total / 3, 2),
        "payroll_total":        payroll_total,
        "payroll_annual_avg":   round(payroll_total / 3, 2),
        "amortization_total":   amortization_total,
        "amortization_annual_avg": round(amortization_total / 3, 2),
    }


# ---------------------------------------------------------------------------
# Step 3: Read the shareholder files
# ---------------------------------------------------------------------------

def parse_shareholder_sheet(rows, person_label, acct_key):
    """
    Extract opening balance, closing balance, and all transactions
    for a given sub-account (e.g. '2901 Ramzan Khuwaja').

    Row structure (from QuickBooks export):
        col 0: None or sub-account label or section label
        col 1: account name or 'Beginning Balance' or 'TOTAL'
        col 2: transaction date (DD/MM/YYYY string)
        col 3: transaction type
        col 4: #
        col 5: Name
        col 6: Memo/Description
        col 7: Account full name
        col 8: Cleared
        col 9: Amount
        col 10: Balance
    """
    in_section = False
    opening_balance = 0.0
    closing_balance = 0.0
    transactions = []

    for row in rows:
        # Detect section start
        if row[0] == person_label:
            in_section = True
            continue

        if not in_section:
            continue

        # Detect section end (next sibling or total row)
        label0 = str(row[0]) if row[0] is not None else ""
        label1 = str(row[1]) if row[1] is not None else ""

        # Stop at sibling sub-account or totals
        if row[0] is not None and row[0] != person_label and label0.startswith("29"):
            break
        if label0.startswith("Total for 2900") or label0 == "TOTAL" or label1 == "TOTAL":
            break

        # Beginning balance row
        if label1 == "Beginning Balance":
            opening_balance = safe_float(row[10])
            continue

        # Total for this person
        if label0.startswith(f"Total for {person_label}"):
            # closing balance is the last transaction balance
            break

        # Transaction row: col 1 is the account key (e.g. '2901 Ramzan Khuwaja')
        if label1 == acct_key and row[2] is not None:
            date_str  = str(row[2]) if row[2] is not None else ""
            memo      = str(row[6]) if row[6] is not None else ""
            name      = str(row[5]) if row[5] is not None else ""
            tx_type   = str(row[3]) if row[3] is not None else ""
            tx_num    = str(row[4]) if row[4] is not None else ""
            amount    = safe_float(row[9])
            balance   = safe_float(row[10])
            closing_balance = balance  # updated each row; last one is closing

            transactions.append({
                "date":    date_str,
                "type":    tx_type,
                "num":     tx_num,
                "name":    name,
                "memo":    memo,
                "amount":  round(amount, 2),
                "balance": round(balance, 2),
            })

    return {
        "opening_balance": round(opening_balance, 2),
        "closing_balance": round(closing_balance, 2),
        "transactions":    transactions,
    }


def read_shareholder_file(path, label):
    print(f"Reading {label}...  ({os.path.basename(path)})")
    rows = load_file(path)

    ramzan = parse_shareholder_sheet(rows, "2901 Ramzan Khuwaja", "2901 Ramzan Khuwaja")
    rezai  = parse_shareholder_sheet(rows, "2902 Mohammad Rezai",  "2902 Mohammad Rezai")

    return ramzan, rezai


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("som_mae_financials — extract_data.py")
    print(f"Run at: {datetime.now().isoformat()}")
    print("=" * 60)

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # -----------------------------------------------------------------------
    # 0. Locate source files (tolerates any prefix and .xlsx or .csv)
    # -----------------------------------------------------------------------
    pl_compare_path = find_file("Profit and Loss - Compare YTD")
    pl_3yr_path     = find_file("Profit and Loss - Aug 2022")
    sh_current_path = find_file("Shareholder Advances - this fiscal year")
    sh_all_path     = find_file("Shareholder Advances - all dates")
    print(f"  P&L Compare   : {os.path.basename(pl_compare_path)}")
    print(f"  P&L 3-year    : {os.path.basename(pl_3yr_path)}")
    print(f"  SH Current    : {os.path.basename(sh_current_path)}")
    print(f"  SH All dates  : {os.path.basename(sh_all_path)}")

    # -----------------------------------------------------------------------
    # 1. YTD Compare P&L
    # -----------------------------------------------------------------------
    pl = read_pl_compare(pl_compare_path)

    ytd_cutoff_date      = pl["ytd_cutoff_date"]
    ytd_tuition_current  = pl["ytd_tuition_current"]
    ytd_tuition_prior    = pl["ytd_tuition_prior"]
    qb_profit            = pl["qb_profit"]
    taxes_paid_8500      = pl["taxes_paid_8500"]
    canada_carbon_rebate = pl["canada_carbon_rebate"]

    # -----------------------------------------------------------------------
    # 2. 3-year benchmarks
    # -----------------------------------------------------------------------
    benchmarks = read_pl_3yr(pl_3yr_path)

    # -----------------------------------------------------------------------
    # 3. Shareholder files
    # -----------------------------------------------------------------------
    sh_cy_ramzan, sh_cy_rezai = read_shareholder_file(sh_current_path, "Current-Year Shareholder")
    sh_all_ramzan, sh_all_rezai = read_shareholder_file(sh_all_path, "All-Dates Shareholder")

    # -----------------------------------------------------------------------
    # 4. Derived / computed fields
    # -----------------------------------------------------------------------

    # Pre-tax income proxy (H1): QB profit + taxes + remove carbon rebate
    h1_pretax_proxy = qb_profit + taxes_paid_8500 - canada_carbon_rebate

    # YTD-to-annual ratio: use prior year YTD vs known audited full year
    ytd_to_annual_ratio = (ytd_tuition_prior / PY_FULL_YEAR_TUITION) if PY_FULL_YEAR_TUITION else None

    # Projected full year revenue (current year)
    projected_full_year = (ytd_tuition_current / ytd_to_annual_ratio) if ytd_to_annual_ratio else None

    # YoY growth %
    yoy_growth_pct = change_pct(ytd_tuition_current, ytd_tuition_prior)

    # Marketing obligations
    obligation_conservative = round(ytd_tuition_current * MARKETING_OBLIGATION_RATE, 2)
    obligation_projected    = round(projected_full_year * MARKETING_OBLIGATION_RATE, 2) if projected_full_year else None
    gap_conservative        = round(obligation_conservative - pl["total_marketing_cy"], 2)
    gap_projected           = round(obligation_projected - pl["total_marketing_cy"], 2) if obligation_projected else None

    # Installments paid YTD (Oct 2025 + Jan 2026 = $1,530 + $13,565)
    installments_paid_ytd = 1530.00 + 13565.00

    # Rezai closing balance — round near-zero floats to 0
    rezai_cy_close = sh_cy_rezai["closing_balance"]
    if abs(rezai_cy_close) < 0.01:
        rezai_cy_close = 0.00
    rezai_all_close = sh_all_rezai["closing_balance"]
    if abs(rezai_all_close) < 0.01:
        rezai_all_close = 0.00

    ramzan_close = sh_cy_ramzan["closing_balance"]
    combined_closing = round(ramzan_close + rezai_cy_close, 2)

    # -----------------------------------------------------------------------
    # 5. Build JSON
    # -----------------------------------------------------------------------
    run_data = {
        "meta": {
            "extracted_at":       datetime.now().isoformat(),
            "ytd_cutoff_date":    ytd_cutoff_date,
            "fiscal_year_label":  "August 1, 2025 \u2013 July 31, 2026",
            "fy_label":           "FY2025-26",
        },
        "revenue": {
            "ytd_tuition_current":   round(ytd_tuition_current, 2),
            "ytd_tuition_prior_year": round(ytd_tuition_prior, 2),
            "full_year_prior":        PY_FULL_YEAR_TUITION,
            "ytd_to_annual_ratio":    round(ytd_to_annual_ratio, 6) if ytd_to_annual_ratio else None,
            "projected_full_year":    round(projected_full_year, 2) if projected_full_year else None,
            "yoy_growth_pct":         yoy_growth_pct,
        },
        "income": {
            "qb_profit":             round(qb_profit, 2),
            "taxes_paid_8500":       round(taxes_paid_8500, 2),
            "canada_carbon_rebate":  round(canada_carbon_rebate, 2),
            "h1_pretax_proxy":       round(h1_pretax_proxy, 2),
        },
        "marketing": {
            "total_ytd_current": round(pl["total_marketing_cy"], 2),
            "total_ytd_prior":   round(pl["total_marketing_py"], 2),
            "accounts":          {k: {"current": round(v["current"], 2), "prior": round(v["prior"], 2)}
                                  for k, v in pl["marketing_accounts"].items()},
            "obligation_conservative": obligation_conservative,
            "obligation_projected":    obligation_projected,
            "gap_conservative":        gap_conservative,
            "gap_projected":           gap_projected,
        },
        "expenses": pl["expenses"],
        "benchmarks_3yr": {
            "tuition_total":           round(benchmarks["tuition_total"], 2),
            "tuition_annual_avg":      benchmarks["tuition_annual_avg"],
            "marketing_total":         round(benchmarks["marketing_total"], 2),
            "marketing_annual_avg":    benchmarks["marketing_annual_avg"],
            "ftc_total":               round(benchmarks["ftc_total"], 2),
            "ftc_annual_avg":          benchmarks["ftc_annual_avg"],
            "it_total":                round(benchmarks["it_total"], 2),
            "it_annual_avg":           benchmarks["it_annual_avg"],
            "payroll_total":           round(benchmarks["payroll_total"], 2),
            "payroll_annual_avg":      benchmarks["payroll_annual_avg"],
            "amortization_total":      round(benchmarks["amortization_total"], 2),
            "amortization_annual_avg": benchmarks["amortization_annual_avg"],
        },
        "tax": {
            "fy2024_25": {
                "total_revenue":       3_020_723.00,
                "accounting_income":   310_894.00,
                "taxable_income":      308_658.00,
                "federal_tax":         31_350.00,
                "ontario_tax":         10_876.00,
                "total_tax":           42_226.00,
                "effective_rate":      0.1368,
                "sbd_limit":           300_000.00,
                "cca_total":           87_068.00,
            },
            "installments": [
                {"due": "October 31, 2025",  "amount": 1530.00,  "status": "paid"},
                {"due": "January 31, 2026",  "amount": 13565.00, "status": "paid"},
                {"due": "April 30, 2026",    "amount": 13565.00, "status": "upcoming"},
                {"due": "July 31, 2026",     "amount": 13566.00, "status": "future"},
            ],
            "installments_paid_ytd": installments_paid_ytd,
            "installments_total":    42226.00,
        },
        "shareholder": {
            "ramzan": {
                "opening_balance": sh_cy_ramzan["opening_balance"],
                "closing_balance": round(ramzan_close, 2),
                "transactions":    sh_cy_ramzan["transactions"],
                "all_dates": {
                    "opening_balance": sh_all_ramzan["opening_balance"],
                    "closing_balance": round(sh_all_ramzan["closing_balance"], 2),
                    "transactions":    sh_all_ramzan["transactions"],
                },
            },
            "rezai": {
                "opening_balance": sh_cy_rezai["opening_balance"],
                "closing_balance": rezai_cy_close,
                "transactions":    sh_cy_rezai["transactions"],
                "all_dates": {
                    "opening_balance": sh_all_rezai["opening_balance"],
                    "closing_balance": rezai_all_close,
                    "transactions":    sh_all_rezai["transactions"],
                },
            },
            "combined_closing": combined_closing,
        },
    }

    # -----------------------------------------------------------------------
    # 6. Write JSON
    # -----------------------------------------------------------------------
    print(f"\nWriting run_data.json...  ({OUTPUT_FILE})")
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(run_data, f, indent=2, ensure_ascii=False)
    print("Done.")

    # -----------------------------------------------------------------------
    # 7. Verification output
    # -----------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("VERIFICATION — key values extracted")
    print("=" * 60)
    print(f"  YTD cutoff date           : {ytd_cutoff_date}")
    print(f"  YTD tuition (CY)          : ${ytd_tuition_current:,.2f}")
    print(f"  YTD tuition (PY)          : ${ytd_tuition_prior:,.2f}")
    print(f"  QB Profit                 : ${qb_profit:,.2f}")
    print(f"  Taxes paid (8500)         : ${taxes_paid_8500:,.2f}")
    print(f"  Canada Carbon Rebate      : ${canada_carbon_rebate:,.2f}")
    print(f"  H1 pre-tax proxy          : ${h1_pretax_proxy:,.2f}")
    print(f"  Marketing total (CY)      : ${pl['total_marketing_cy']:,.2f}")
    print(f"  Ramzan closing balance    : ${ramzan_close:,.2f}")
    print(f"  Rezai closing balance     : ${rezai_cy_close:,.2f}")
    print(f"  YTD/Annual ratio          : {ytd_to_annual_ratio:.4%}")
    print(f"  Projected full year rev   : ${projected_full_year:,.2f}")
    print(f"  Marketing obligation (YTD): ${obligation_conservative:,.2f}")
    print(f"  Marketing obligation (proj): ${obligation_projected:,.2f}")
    print(f"  Marketing gap (YTD basis) : ${gap_conservative:,.2f}")
    print(f"  Marketing gap (projected) : ${gap_projected:,.2f}")
    print("=" * 60)

    # Print JSON to stdout as well
    print("\n--- run_data.json ---")
    print(json.dumps(run_data, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
